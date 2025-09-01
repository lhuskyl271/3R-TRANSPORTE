# ventas/views.py

import boto3
import os
from botocore.exceptions import BotoCoreError, NoCredentialsError
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
)
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy, reverse
from .models import (
    Prospecto, Interaccion, Recordatorio, Etiqueta, Trabajador, 
    ProspectoTrabajador, ArchivoAdjunto, Proyecto, Entregable, 
    EquipoProyecto, SeguimientoProyecto,KanbanColumna, KanbanTarea,DiagramaProyecto 
)
from .forms import (
    ProspectoForm, InteraccionForm, RecordatorioForm, TrabajadorForm, 
    ProspectoTrabajadorForm, ProspectoTrabajadorUpdateForm, ArchivoAdjuntoForm,
    ProyectoUpdateForm, AsignarMiembroEquipoForm, EntregableForm, SeguimientoProyectoForm, KanbanTareaForm # <-- Nuevos
)
from django.db.models import Count, Q, Avg
from django.http import HttpResponseForbidden, HttpResponse
from openpyxl import Workbook
from django.contrib import messages
from django.db.models.functions import ExtractDay, Now
from openpyxl.styles import Font
from django.utils import timezone
import json
from datetime import timedelta
import pytz 
from django.core.paginator import Paginator
from django.db.models import Subquery, OuterRef, Case, When, F, IntegerField
from django.http import JsonResponse
from django.views.generic import TemplateView #


# ==============================================================================
# MIXINS Y VISTAS BASE
# ==============================================================================

class OwnerRequiredMixin:
    """
    Mixin para asegurar que solo el superusuario o el usuario asignado
    puedan modificar objetos relacionados a un prospecto.
    """
    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        
        asignado = None
        if isinstance(obj, Prospecto):
            asignado = obj.asignado_a
        elif isinstance(obj, (Interaccion, Recordatorio, ProspectoTrabajador, ArchivoAdjunto)):
            asignado = obj.prospecto.asignado_a
        
        if asignado and asignado != self.request.user and not self.request.user.is_superuser:
            raise HttpResponseForbidden("No tienes permiso para realizar esta acción.")
        return obj

# ==============================================================================
# VISTAS DEL DASHBOARD Y PROSPECTOS
# ==============================================================================

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'ventas/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        try:
            user_timezone = pytz.timezone('America/Mexico_City') 
        except pytz.UnknownTimeZoneError:
            user_timezone = pytz.timezone(timezone.get_default_timezone_name())

        hoy = timezone.now().astimezone(user_timezone)
        
        prospectos_qs = Prospecto.objects.all()
        if not user.is_superuser:
            prospectos_qs = prospectos_qs.filter(asignado_a=user)

        context['total_prospectos'] = prospectos_qs.count()
        
        quince_dias_atras = hoy - timedelta(days=15)
        context['prospectos_nuevos'] = prospectos_qs.filter(
            estado=Prospecto.Estado.NUEVO,
            fecha_creacion__gte=quince_dias_atras
        ).count()
        
        context['clientes_ganados'] = prospectos_qs.filter(estado=Prospecto.Estado.GANADO).count()
        
        reporte_data = prospectos_qs.values('estado').annotate(total=Count('estado')).order_by('estado')
        estado_display_map = dict(Prospecto.Estado.choices) 
        chart_data = {
            "labels": [estado_display_map.get(item['estado'], item['estado']) for item in reporte_data],
            "data": [item['total'] for item in reporte_data],
        }
        context['chart_data_json'] = json.dumps(chart_data)
        
        promedio_calificaciones = ProspectoTrabajador.objects.filter(
            prospecto__in=prospectos_qs
        ).values('trabajador__nombre').annotate(promedio=Avg('calificacion')).order_by('-promedio')
        context['promedio_calificaciones_trabajador'] = promedio_calificaciones

        ultima_interaccion_subquery = Interaccion.objects.filter(
            prospecto=OuterRef('pk')
        ).order_by('-fecha').values('fecha')[:1]
        
        prospectos_inactivos = prospectos_qs.exclude(
            estado__in=[Prospecto.Estado.GANADO, Prospecto.Estado.PERDIDO]
        ).annotate(
            ultima_interaccion=Subquery(ultima_interaccion_subquery)
        ).annotate(
            dias_inactivo=Case(
                When(
                    ultima_interaccion__isnull=True, 
                    then=ExtractDay(Now() - F('fecha_creacion'))
                ),
                When(
                    ultima_interaccion__isnull=False,
                    then=ExtractDay(Now() - F('ultima_interaccion'))
                ),
                output_field=IntegerField()
            )
        ).filter(
            dias_inactivo__gte=1
        ).order_by('-dias_inactivo')
        
        page_size = int(self.request.GET.get('page_size', 10))
        paginator = Paginator(prospectos_inactivos, page_size)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context['prospectos_inactivos'] = page_obj
        context['seguimiento_requerido_count'] = prospectos_inactivos.count()

        quince_dias_despues = hoy + timedelta(days=15)
        recordatorios_proximos = Recordatorio.objects.filter(
            prospecto__in=prospectos_qs, completado=False, 
            fecha_recordatorio__gte=hoy, fecha_recordatorio__lte=quince_dias_despues
        ).select_related('prospecto').order_by('fecha_recordatorio')
        context['recordatorios_proximos'] = recordatorios_proximos
        
        recordatorios_pasados = Recordatorio.objects.filter(
            prospecto__in=prospectos_qs, 
            completado=False, 
            fecha_recordatorio__lt=hoy
        ).select_related('prospecto').order_by('fecha_recordatorio')
        context['recordatorios_pasados'] = recordatorios_pasados
        
        return context


class ProspectoListView(LoginRequiredMixin, ListView):
    model = Prospecto
    template_name = 'ventas/prospecto_list.html'
    context_object_name = 'prospectos'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            queryset = queryset.filter(asignado_a=self.request.user)

        estado_filter = self.request.GET.get('estado')
        if estado_filter:
            queryset = queryset.filter(estado=estado_filter)
        
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(nombre_completo__icontains=query) |
                Q(email__icontains=query) |
                Q(empresa__icontains=query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = self.model.objects.all()
        if not self.request.user.is_superuser:
            base_qs = base_qs.filter(asignado_a=self.request.user)

        status_counts_dict = {
            item['estado']: item['total'] 
            for item in base_qs.values('estado').annotate(total=Count('id'))
        }

        status_cards_data = []
        for value, name in Prospecto.Estado.choices:
            status_cards_data.append({
                'value': value,
                'name': name,
                'count': status_counts_dict.get(value, 0)
            })

        context['status_cards'] = status_cards_data
        context['total_prospectos_global'] = base_qs.count()
        return context

class ProspectoDetailView(LoginRequiredMixin, OwnerRequiredMixin, DetailView):
    model = Prospecto
    template_name = 'ventas/prospecto_detail.html'
    context_object_name = 'prospecto'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        prospecto = self.get_object() # <-- Importante obtener el objeto prospecto

        # --- Lógica estándar que ya tenías ---
        context['interaccion_form'] = InteraccionForm()
        context['recordatorio_form'] = RecordatorioForm()
        context['trabajador_form'] = ProspectoTrabajadorForm()
        context['archivo_form'] = ArchivoAdjuntoForm() 
        context['archivos_adjuntos'] = prospecto.archivos_adjuntos.all()
        
        trabajadores_asociados_ids = prospecto.trabajadores.values_list('id', flat=True)
        context['trabajador_form'].fields['trabajador'].queryset = Trabajador.objects.exclude(id__in=trabajadores_asociados_ids)
        
        context['relaciones_trabajadores'] = prospecto.prospectotrabajador_set.all().select_related('trabajador')
        context['interacciones'] = prospecto.interacciones.all().order_by('-fecha')
        context['recordatorios'] = prospecto.recordatorios.all().order_by('completado', 'fecha_recordatorio')

        # --- ✅ LÓGICA FALTANTE PARA GESTIÓN DE PROYECTO ---
        # Si el prospecto es un cliente ganado, obtenemos o creamos su proyecto.
        if prospecto.estado == Prospecto.Estado.GANADO:
            proyecto, created = Proyecto.objects.get_or_create(prospecto=prospecto)
            if created and not proyecto.nombre_proyecto:
                # Si se acaba de crear, le damos un nombre por defecto
                proyecto.nombre_proyecto = f"Proyecto para {prospecto.empresa or prospecto.nombre_completo}"
                proyecto.save()

            # Se añade el proyecto y los formularios al contexto
            context['proyecto'] = proyecto
            context['proyecto_form'] = ProyectoUpdateForm(instance=proyecto)
            context['entregable_form'] = EntregableForm()
            context['seguimiento_form'] = SeguimientoProyectoForm()
            context['asignar_miembro_form'] = AsignarMiembroEquipoForm()
            
            # Se añaden los datos relacionados al proyecto
            context['equipo_proyecto'] = proyecto.equipoproyecto_set.all().select_related('trabajador')
            context['entregables'] = proyecto.entregables.all()
            context['seguimientos'] = proyecto.seguimientos.all().select_related('creado_por')

        return context
    
class ProspectoCreateView(LoginRequiredMixin, CreateView):
    model = Prospecto
    form_class = ProspectoForm
    template_name = 'ventas/prospecto_form.html'
    
    def form_valid(self, form):
        form.instance.asignado_a = self.request.user
        messages.success(self.request, f"Prospecto '{form.instance.nombre_completo}' creado exitosamente.")
        return super().form_valid(form)

class ProspectoUpdateView(LoginRequiredMixin, OwnerRequiredMixin, UpdateView):
    model = Prospecto
    form_class = ProspectoForm
    template_name = 'ventas/prospecto_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, f"Prospecto '{self.object.nombre_completo}' actualizado correctamente.")
        return super().form_valid(form)

class ProspectoDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    model = Prospecto
    template_name = 'ventas/prospecto_confirm_delete.html'
    success_url = reverse_lazy('prospecto-list')

    def form_valid(self, form):
        messages.success(self.request, f"Prospecto '{self.object.nombre_completo}' ha sido eliminado.")
        return super().form_valid(form)

class TrabajadorListView(LoginRequiredMixin, ListView):
    model = Trabajador
    template_name = 'ventas/trabajador_list.html'
    context_object_name = 'trabajadores'
    paginate_by = 15

class TrabajadorCreateView(LoginRequiredMixin, CreateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'ventas/trabajador_form.html'
    success_url = reverse_lazy('trabajador-list')

class TrabajadorUpdateView(LoginRequiredMixin, UpdateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'ventas/trabajador_form.html'
    success_url = reverse_lazy('trabajador-list')

class TrabajadorDeleteView(LoginRequiredMixin, DeleteView):
    model = Trabajador
    template_name = 'ventas/trabajador_confirm_delete.html'
    success_url = reverse_lazy('trabajador-list')

@login_required
def add_trabajador_a_prospecto(request, prospecto_pk):
    prospecto = get_object_or_404(Prospecto, pk=prospecto_pk)
    if request.method == 'POST':
        form = ProspectoTrabajadorForm(request.POST)
        if form.is_valid():
            trabajador = form.cleaned_data['trabajador']
            if ProspectoTrabajador.objects.filter(prospecto=prospecto, trabajador=trabajador).exists():
                messages.error(request, f"El empleado '{trabajador.nombre}' ya está asignado a este prospecto.")
            else:
                relacion = form.save(commit=False)
                relacion.prospecto = prospecto
                relacion.save()
                messages.success(request, f"¡El empleado '{relacion.trabajador.nombre}' fue asignado correctamente!")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en el campo '{form.fields[field].label}': {error}")
    return redirect('prospecto-detail', pk=prospecto_pk)

class ProspectoTrabajadorUpdateView(LoginRequiredMixin, OwnerRequiredMixin, UpdateView):
    model = ProspectoTrabajador
    form_class = ProspectoTrabajadorUpdateForm
    template_name = 'ventas/prospecto_trabajador_form.html'

    def get_success_url(self):
        messages.success(self.request, f"Se actualizó la calificación para '{self.object.trabajador.nombre}'.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

class ProspectoTrabajadorDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    model = ProspectoTrabajador
    template_name = 'ventas/prospecto_trabajador_confirm_delete.html'
    
    def get_success_url(self):
        messages.success(self.request, f"Se eliminó la relación con '{self.object.trabajador.nombre}'.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

@login_required
def add_interaccion(request, prospecto_pk):
    prospecto = get_object_or_404(Prospecto, pk=prospecto_pk)
    if request.method == 'POST':
        form = InteraccionForm(request.POST)
        if form.is_valid():
            interaccion = form.save(commit=False)
            interaccion.prospecto = prospecto
            interaccion.creado_por = request.user
            interaccion.save()
            messages.success(request, "Interacción registrada.")
    return redirect('prospecto-detail', pk=prospecto_pk)

class InteraccionUpdateView(LoginRequiredMixin, OwnerRequiredMixin, UpdateView):
    model = Interaccion
    form_class = InteraccionForm
    template_name = 'ventas/interaccion_form.html'
    def get_success_url(self):
        messages.success(self.request, "Interacción actualizada.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

class InteraccionDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    model = Interaccion
    template_name = 'ventas/interaccion_confirm_delete.html'
    def get_success_url(self):
        messages.success(self.request, "Interacción eliminada.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

@login_required
def add_recordatorio(request, prospecto_pk):
    prospecto = get_object_or_404(Prospecto, pk=prospecto_pk)
    if request.method == 'POST':
        form = RecordatorioForm(request.POST)
        if form.is_valid():
            recordatorio = form.save(commit=False)
            recordatorio.prospecto = prospecto
            recordatorio.creado_por = request.user
            recordatorio.save()
            messages.success(request, "Recordatorio creado.")
    return redirect('prospecto-detail', pk=prospecto_pk)

class RecordatorioUpdateView(LoginRequiredMixin, OwnerRequiredMixin, UpdateView):
    model = Recordatorio
    form_class = RecordatorioForm
    template_name = 'ventas/recordatorio_form.html'
    def get_success_url(self):
        messages.success(self.request, "Recordatorio actualizado.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

class RecordatorioDeleteView(LoginRequiredMixin, OwnerRequiredMixin, DeleteView):
    model = Recordatorio
    template_name = 'ventas/recordatorio_confirm_delete.html'
    def get_success_url(self):
        messages.success(self.request, "Recordatorio eliminado.")
        return reverse('prospecto-detail', kwargs={'pk': self.object.prospecto.pk})

@login_required
def toggle_recordatorio(request, pk):
    recordatorio = get_object_or_404(Recordatorio, pk=pk)
    if recordatorio.prospecto.asignado_a != request.user and not request.user.is_superuser:
        return HttpResponseForbidden("No tienes permiso.")
    
    recordatorio.completado = not recordatorio.completado
    recordatorio.save()
    status = "completado" if recordatorio.completado else "marcado como pendiente"
    messages.info(request, f"Recordatorio '{recordatorio.titulo}' {status}.")
    return redirect('prospecto-detail', pk=recordatorio.prospecto.pk)

@login_required
def export_prospectos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = timezone.now().strftime('%Y-%m-%d_%H-%M')
    response['Content-Disposition'] = f'attachment; filename="prospectos_{timestamp}.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Prospectos'
    
    headers = [
        'Nombre Completo', 'Email', 'Teléfono', 'Empresa', 'Puesto', 'Estado', 
        'Interés', 'Calificación Prom.', 'Referido Por', 'Contacto Ref.',
        'Detalle Interés', 'Trabajadores', 'Etiquetas', 'Asignado a', 'Fecha Creación'
    ]
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.font = Font(bold=True)

    prospectos_qs = Prospecto.objects.annotate(
        promedio_calificacion=Avg('prospectotrabajador__calificacion')
    )
    if not request.user.is_superuser:
        prospectos_qs = prospectos_qs.filter(asignado_a=request.user)
    
    prospectos = prospectos_qs.select_related('asignado_a').prefetch_related('etiquetas', 'trabajadores')

    for row_num, prospecto in enumerate(prospectos, 2):
        calificacion_str = f"{prospecto.promedio_calificacion:.2f}" if prospecto.promedio_calificacion else "N/A"
        row_data = [
            prospecto.nombre_completo, prospecto.email, prospecto.telefono, prospecto.empresa, prospecto.puesto,
            prospecto.get_estado_display(), prospecto.get_interes_principal_display(), calificacion_str,
            prospecto.referencio, prospecto.contacto_referencio, prospecto.interes_cliente,
            ", ".join([t.nombre for t in prospecto.trabajadores.all()]),
            ", ".join([e.nombre for e in prospecto.etiquetas.all()]),
            prospecto.asignado_a.username if prospecto.asignado_a else '',
            # Se cambió 'fecha_creation' por 'fecha_creacion'
            prospecto.fecha_creacion.strftime('%Y-%m-%d %H:%M') if prospecto.fecha_creacion else ''
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    for i, column_cells in enumerate(worksheet.columns):
        try:
            max_length = 0
            column = chr(65 + i)
            for cell in column_cells:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        except:
            pass
            
    workbook.save(response)
    return response

@login_required
def add_archivo(request, prospecto_pk):
    """
    Gestiona la subida de un archivo a S3 usando Boto3 directamente y lo asocia
    con un prospecto específico. El nombre del archivo se toma automáticamente.
    """
    prospecto = get_object_or_404(Prospecto, pk=prospecto_pk)
    if request.method != 'POST':
        return HttpResponse("This view only accepts POST requests.", status=405)

    form = ArchivoAdjuntoForm(request.POST, request.FILES)
    if form.is_valid():
        uploaded_file = form.cleaned_data['archivo']
        titulo_archivo = uploaded_file.name

        # --- CORRECCIÓN IMPORTANTE ---
        # Construimos la ruta RELATIVA, SIN el prefijo 'media/'.
        # Django-storages lo añadirá automáticamente al generar la URL.
        # ej: prospectos/7/documento_propuesta.pdf
        s3_key = f"prospectos/{prospecto.pk}/{titulo_archivo}"

        try:
            s3_client = boto3.client(
                's3',
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_S3_REGION_NAME
            )

            # La ruta de subida necesita el prefijo 'media/' porque Boto3 no lo conoce.
            full_s3_path = f"{settings.AWS_LOCATION}/{s3_key}"

            s3_client.upload_fileobj(
                uploaded_file,
                settings.AWS_STORAGE_BUCKET_NAME,
                full_s3_path
            )
            
            # Guardamos en la base de datos la ruta SIN el prefijo 'media/'.
            archivo_adjunto = ArchivoAdjunto(
                prospecto=prospecto,
                nombre=titulo_archivo,
                archivo=s3_key # <-- Se guarda la ruta relativa
            )
            archivo_adjunto.save()

            messages.success(request, f"Archivo '{titulo_archivo}' subido exitosamente.")

        except (BotoCoreError, NoCredentialsError) as e:
            messages.error(request, f"Error de configuración o conexión con S3: {e}")
        except Exception as e:
            messages.error(request, f"Ocurrió un error inesperado al subir el archivo: {e}")

    else:
        error_string = " ".join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()])
        messages.error(request, f"Error en el formulario. Detalles: {error_string}")

    return redirect('prospecto-detail', pk=prospecto_pk)

@login_required
def delete_archivo(request, pk):
    """
    Elimina un archivo adjunto tanto de S3 (usando Boto3) como de la base de datos.
    """
    archivo = get_object_or_404(ArchivoAdjunto, pk=pk)
    
    if archivo.prospecto.asignado_a != request.user and not request.user.is_superuser:
        messages.error(request, "No tienes permiso para eliminar este archivo.")
        return redirect('prospecto-detail', pk=archivo.prospecto.pk)
    
    prospecto_pk = archivo.prospecto.pk
    file_name = archivo.nombre
    
    # --- CORRECCIÓN IMPORTANTE ---
    # La ruta completa en S3 incluye el prefijo 'media/'
    full_s3_path = archivo.archivo.name

    try:
        s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_S3_REGION_NAME
        )
        
        s3_client.delete_object(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=full_s3_path
        )

        archivo.delete()
        
        messages.success(request, f"El archivo '{file_name}' ha sido eliminado exitosamente.")

    except (BotoCoreError, NoCredentialsError) as e:
        messages.error(request, f"Error de conexión con S3 al intentar borrar: {e}")
    except Exception as e:
        messages.error(request, f"No se pudo eliminar el archivo del servidor: {e}")

    return redirect('prospecto-detail', pk=prospecto_pk)

class CalendarioView(LoginRequiredMixin, TemplateView):
    """
    Renderiza la página principal que contendrá el calendario.
    """
    template_name = 'ventas/calendario.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = "Calendario de Actividades"
        return context

@login_required
def calendario_eventos(request):
    """
    Proporciona los eventos (recordatorios) en formato JSON para FullCalendar.
    """
    user = request.user
    
    # Filtrar recordatorios basados en el usuario (superuser ve todo)
    if user.is_superuser:
        recordatorios = Recordatorio.objects.all()
    else:
        recordatorios = Recordatorio.objects.filter(prospecto__asignado_a=user)
        
    eventos = []
    for recordatorio in recordatorios:
        # Asignar un color basado en el estado del recordatorio
        color = '#2ecc71' if recordatorio.completado else '#e74c3c' # Verde si está completado, rojo si no

        eventos.append({
            'title': f"{recordatorio.titulo} ({recordatorio.prospecto.nombre_completo})",
            'start': recordatorio.fecha_recordatorio.isoformat(),
            'end': recordatorio.fecha_recordatorio.isoformat(),
            'url': reverse('prospecto-detail', kwargs={'pk': recordatorio.prospecto.pk}),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'description': f"Prospecto: {recordatorio.prospecto.nombre_completo}",
                'status': 'Completado' if recordatorio.completado else 'Pendiente'
            }
        })
        
    return JsonResponse(eventos, safe=False)

class ClienteCerradoListView(LoginRequiredMixin, ListView):
    """
    Vista para listar únicamente los prospectos que han sido marcados como 'GANADO'.
    """
    model = Prospecto
    template_name = 'ventas/cliente_cerrado_list.html' # Usamos una nueva plantilla
    context_object_name = 'clientes'
    paginate_by = 10

    def get_queryset(self):
        # Filtramos para obtener solo prospectos con estado 'GANADO'
        queryset = super().get_queryset().filter(estado=Prospecto.Estado.GANADO)
        
        # Si el usuario no es superusuario, filtramos por sus propios clientes
        if not self.request.user.is_superuser:
            queryset = queryset.filter(asignado_a=self.request.user)

        # Mantenemos la funcionalidad de búsqueda
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(nombre_completo__icontains=query) |
                Q(email__icontains=query) |
                Q(empresa__icontains=query)
            )
        
        return queryset.order_by('-fecha_actualizacion')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Añadimos un título personalizado para la plantilla
        context['page_title'] = 'Clientes Cerrados'
        return context
    
@login_required
def update_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    # Aquí puedes añadir validación de permisos si es necesario
    if request.method == 'POST':
        form = ProyectoUpdateForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, "Los detalles del proyecto han sido actualizados.")
        else:
            messages.error(request, "Hubo un error al actualizar el proyecto.")
    return redirect('prospecto-detail', pk=proyecto.prospecto.pk)


@login_required
def add_entregable(request, proyecto_pk):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_pk)
    if request.method == 'POST':
        form = EntregableForm(request.POST)
        if form.is_valid():
            entregable = form.save(commit=False)
            entregable.proyecto = proyecto
            entregable.save()
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string('ventas/snippets/entregable_item.html', {'entregable': entregable}, request=request)
                return JsonResponse({
                    'status': 'success',
                    'message': 'Entregable añadido correctamente.',
                    'action': 'create',
                    'list_id': 'entregables-list',
                    'html': html,
                    'deliverable_count': proyecto.entregables.count()
                })
            messages.success(request, "Entregable añadido correctamente.")
            return redirect('prospecto-detail', pk=proyecto.prospecto.pk)
    
    # Manejo de error si el formulario no es válido
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'error', 'message': 'Hubo un error con los datos enviados.'}, status=400)
    messages.error(request, "Error al añadir el entregable.")
    return redirect('prospecto-detail', pk=proyecto.prospecto.pk)

@login_required
def add_seguimiento_proyecto(request, proyecto_pk):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_pk)
    if request.method == 'POST':
        form = SeguimientoProyectoForm(request.POST)
        if form.is_valid():
            seguimiento = form.save(commit=False)
            seguimiento.proyecto = proyecto
            seguimiento.creado_por = request.user
            seguimiento.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string('ventas/snippets/seguimiento_item.html', {'item': seguimiento}, request=request)
                return JsonResponse({
                    'status': 'success',
                    'message': 'Seguimiento del proyecto registrado.',
                    'action': 'create',
                    'list_id': 'seguimiento-list',
                    'html': html
                })
            messages.success(request, "Seguimiento del proyecto registrado.")
            return redirect('prospecto-detail', pk=proyecto.prospecto.pk)
            
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'error', 'message': 'La nota no puede estar vacía.'}, status=400)
    messages.error(request, "Error al registrar el seguimiento.")
    return redirect('prospecto-detail', pk=proyecto.prospecto.pk)

@login_required
def asignar_miembro_equipo(request, proyecto_pk):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_pk)
    if request.method == 'POST':
        form = AsignarMiembroEquipoForm(request.POST)
        if form.is_valid():
            miembro = form.save(commit=False)
            if EquipoProyecto.objects.filter(proyecto=proyecto, trabajador=miembro.trabajador).exists():
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'error', 'message': f"El trabajador '{miembro.trabajador}' ya forma parte del equipo."}, status=400)
                messages.warning(request, f"El trabajador '{miembro.trabajador}' ya forma parte del equipo.")
            else:
                miembro.proyecto = proyecto
                miembro.save()
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    html = render_to_string('ventas/snippets/miembro_item.html', {'miembro': miembro}, request=request)
                    return JsonResponse({
                        'status': 'success',
                        'message': f"'{miembro.trabajador}' ha sido añadido al equipo del proyecto.",
                        'action': 'create',
                        'list_id': 'miembros-list',
                        'html': html,
                        'team_count': proyecto.equipoproyecto_set.count()
                    })
                messages.success(request, f"'{miembro.trabajador}' ha sido añadido al equipo del proyecto.")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': 'Error al asignar al miembro del equipo.'}, status=400)
            messages.error(request, "Error al asignar al miembro del equipo.")
    
    return redirect('prospecto-detail', pk=proyecto.prospecto.pk)


class ProyectoDetailView(LoginRequiredMixin, DetailView):
    """
    Vista detallada para la gestión de un proyecto específico.
    Funciona como el dashboard principal del proyecto.
    """
    model = Proyecto
    template_name = 'ventas/proyecto_detail.html'
    context_object_name = 'proyecto'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proyecto = self.get_object()

        # Añadimos los formularios necesarios para las acciones dentro del panel
        context['proyecto_form'] = ProyectoUpdateForm(instance=proyecto)
        context['entregable_form'] = EntregableForm()
        context['seguimiento_form'] = SeguimientoProyectoForm()
        context['asignar_miembro_form'] = AsignarMiembroEquipoForm()

        # Pasamos al contexto toda la información relacionada para mostrarla
        context['equipo_proyecto'] = proyecto.equipoproyecto_set.all().select_related('trabajador')
        context['entregables'] = proyecto.entregables.all()
        context['seguimientos'] = proyecto.seguimientos.all().select_related('creado_por')
        
        # También pasamos el prospecto (cliente) para tener acceso a su información
        context['cliente'] = proyecto.prospecto

        return context
    
class ProyectoFlujoTrabajoView(LoginRequiredMixin, DetailView):
    model = Proyecto
    template_name = 'ventas/proyecto_flujo_trabajo.html'
    context_object_name = 'proyecto'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        proyecto = self.get_object()
        boards = []
        columnas = proyecto.kanban_columnas.prefetch_related('tareas').all()
        
        for columna in columnas:
            items = []
            for tarea in columna.tareas.all():
                items.append({
                    'id': str(tarea.id),
                    'title': tarea.titulo,
                    'description': tarea.descripcion,
                })
            
            boards.append({
                'id': str(columna.id),
                'title': columna.titulo,
                'icon': columna.icono,  # <-- ✅ Pasamos el ícono al frontend
                'item': items
            })
        
        context['kanban_data_json'] = json.dumps(boards)
        return context


@login_required
def mover_tarea_api(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        tarea_id = data.get('tarea_id')
        nueva_columna_id = data.get('nueva_columna_id')
        
        try:
            tarea = KanbanTarea.objects.get(pk=tarea_id)
            nueva_columna = KanbanColumna.objects.get(pk=nueva_columna_id)
            
            tarea.columna = nueva_columna
            # Aquí podrías añadir lógica para reordenar las tareas
            tarea.save()
            
            return JsonResponse({'status': 'success'})
        except (KanbanTarea.DoesNotExist, KanbanColumna.DoesNotExist):
            return JsonResponse({'status': 'error', 'message': 'Tarea o columna no encontrada'}, status=404)
            
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def crear_columna_api(request, proyecto_pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        titulo = data.get('titulo')
        icono = data.get('icono', '')  # <-- ✅ Obtenemos el ícono
        proyecto = get_object_or_404(Proyecto, pk=proyecto_pk)
        
        if titulo:
            ultima_columna = proyecto.kanban_columnas.order_by('-orden').first()
            nuevo_orden = (ultima_columna.orden + 1) if ultima_columna else 0
            
            columna = KanbanColumna.objects.create(
                proyecto=proyecto, 
                titulo=titulo, 
                icono=icono,  # <-- ✅ Guardamos el ícono
                orden=nuevo_orden
            )
            return JsonResponse({'status': 'success', 'id': str(columna.id), 'titulo': columna.titulo, 'icono': columna.icono})
    return JsonResponse({'status': 'error'}, status=400)


# --- NUEVA VISTA API ---
@login_required
def actualizar_columna_api(request, columna_pk):
    columna = get_object_or_404(KanbanColumna, pk=columna_pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        nuevo_titulo = data.get('titulo')
        nuevo_icono = data.get('icono') # Puede ser None si no se envía

        if nuevo_titulo and nuevo_titulo.strip():
            columna.titulo = nuevo_titulo
        
        if nuevo_icono is not None: # Si se envió el campo 'icono' (incluso si está vacío)
            columna.icono = nuevo_icono

        columna.save()
        return JsonResponse({'status': 'success', 'nuevo_titulo': columna.titulo, 'nuevo_icono': columna.icono})
    return JsonResponse({'status': 'error', 'message': 'Petición inválida'}, status=400)
# --- NUEVA VISTA API ---
@login_required
def eliminar_columna_api(request, columna_pk):
    """API para eliminar una columna y todas sus tareas."""
    columna = get_object_or_404(KanbanColumna, pk=columna_pk)
    if request.method == 'POST':
        columna.delete() # Gracias a on_delete=CASCADE, las tareas se borrarán también
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Petición inválida'}, status=400)


@login_required
def crear_tarea_api(request, columna_pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        titulo = data.get('titulo')
        columna = get_object_or_404(KanbanColumna, pk=columna_pk)
        
        if titulo:
            ultima_tarea = columna.tareas.order_by('-orden').first()
            nuevo_orden = (ultima_tarea.orden + 1) if ultima_tarea else 0

            tarea = KanbanTarea.objects.create(columna=columna, titulo=titulo, orden=nuevo_orden)
            return JsonResponse({'status': 'success', 'id': str(tarea.id), 'titulo': tarea.titulo})
    return JsonResponse({'status': 'error', 'message': 'Título no proporcionado'}, status=400)

# --- NUEVA VISTA API ---
@login_required
def actualizar_tarea_api(request, tarea_pk):
    """API para actualizar los detalles de una tarea."""
    tarea = get_object_or_404(KanbanTarea, pk=tarea_pk)
    if request.method == 'POST':
        data = json.loads(request.body)
        # Usamos el form para validar y limpiar los datos
        form = KanbanTareaForm(data, instance=tarea)
        if form.is_valid():
            form.save()
            # Devolvemos los datos actualizados para reflejarlos en el frontend
            return JsonResponse({
                'status': 'success',
                'id': str(tarea.id),
                'titulo': form.cleaned_data['titulo'],
                'descripcion': form.cleaned_data['descripcion']
            })
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Petición inválida'}, status=400)

# --- NUEVA VISTA API ---
@login_required
def eliminar_tarea_api(request, tarea_pk):
    """API para eliminar una tarea."""
    tarea = get_object_or_404(KanbanTarea, pk=tarea_pk)
    if request.method == 'POST':
        tarea.delete()
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Petición inválida'}, status=400)


class EntregableUpdateView(LoginRequiredMixin, UpdateView):
    model = Entregable
    form_class = EntregableForm
    template_name = 'ventas/snippets/entregable_form.html'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            return render(request, self.template_name, self.get_context_data())
        return super().get(request, *args, **kwargs)

    def form_valid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            entregable = form.save()
            html = render_to_string('ventas/snippets/entregable_item.html', {'entregable': entregable}, request=self.request)
            return JsonResponse({
                'status': 'success',
                'message': f"Entregable '{entregable.nombre}' actualizado.",
                'action': 'update',
                'list_id': 'entregables-list',
                'id': entregable.pk,
                'html': html
            })
        messages.success(self.request, f"Entregable '{self.object.nombre}' actualizado.")
        return redirect('prospecto-detail', kwargs={'pk': self.object.proyecto.prospecto.pk})

class EntregableDeleteView(LoginRequiredMixin, DeleteView):
    model = Entregable
    template_name = 'ventas/snippets/entregable_confirm_delete.html'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            return render(request, self.template_name, self.get_context_data())
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            entregable = self.get_object()
            entregable_id = entregable.pk
            entregable_nombre = entregable.nombre
            proyecto = entregable.proyecto
            entregable.delete()
            return JsonResponse({
                'status': 'success',
                'message': f"Entregable '{entregable_nombre}' eliminado.",
                'action': 'delete',
                'list_id': 'entregables-list',
                'id': entregable_id,
                'deliverable_count': proyecto.entregables.count()
            })
        
        self.object = self.get_object()
        success_url = self.get_success_url()
        messages.success(self.request, f"Entregable '{self.object.nombre}' eliminado.")
        self.object.delete()
        return redirect(success_url)
    
class DesasignarMiembroEquipoView(LoginRequiredMixin, DeleteView):
    model = EquipoProyecto
    template_name = 'ventas/snippets/miembro_confirm_delete.html'

    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            return render(request, self.template_name, self.get_context_data())
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            miembro = self.get_object()
            miembro_id = miembro.pk
            trabajador_nombre = miembro.trabajador.nombre
            proyecto = miembro.proyecto
            miembro.delete()
            return JsonResponse({
                'status': 'success',
                'message': f"Se ha quitado a '{trabajador_nombre}' del equipo.",
                'action': 'delete',
                'list_id': 'miembros-list',
                'id': miembro_id,
                'team_count': proyecto.equipoproyecto_set.count()
            })
        
        return super().post(request, *args, **kwargs)
class SeguimientoProyectoUpdateView(LoginRequiredMixin, UpdateView):
    """Actualiza una nota de seguimiento."""
    model = SeguimientoProyecto
    form_class = SeguimientoProyectoForm
    template_name = 'ventas/snippets/seguimiento_form.html'

    # ✅ AÑADIR ESTE MÉTODO
    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            return render(request, self.template_name, self.get_context_data())
        return super().get(request, *args, **kwargs)

    # ✅ AÑADIR LÓGICA AJAX A form_valid
    def form_valid(self, form):
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            seguimiento = form.save()
            html = render_to_string('ventas/snippets/seguimiento_item.html', {'item': seguimiento}, request=self.request)
            return JsonResponse({
                'status': 'success',
                'message': 'La nota de seguimiento ha sido actualizada.',
                'action': 'update',
                'list_id': 'seguimiento-list',
                'id': seguimiento.pk,
                'html': html
            })
        messages.success(self.request, "La nota de seguimiento ha sido actualizada.")
        return redirect('prospecto-detail', kwargs={'pk': self.object.proyecto.prospecto.pk})

class SeguimientoProyectoDeleteView(LoginRequiredMixin, DeleteView):
    """Elimina una nota de seguimiento."""
    model = SeguimientoProyecto
    template_name = 'ventas/snippets/seguimiento_confirm_delete.html'

    # ✅ AÑADIR ESTE MÉTODO
    def get(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            self.object = self.get_object()
            return render(request, self.template_name, self.get_context_data())
        return super().get(request, *args, **kwargs)

    # ✅ REEMPLAZAR EL MÉTODO post
    def post(self, request, *args, **kwargs):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            seguimiento = self.get_object()
            seguimiento_id = seguimiento.pk
            seguimiento.delete()
            return JsonResponse({
                'status': 'success',
                'message': 'La nota de seguimiento ha sido eliminada.',
                'action': 'delete',
                'list_id': 'seguimiento-list',
                'id': seguimiento_id
            })
        
        # Lógica original como fallback
        self.object = self.get_object()
        success_url = self.get_success_url()
        messages.success(self.request, "La nota de seguimiento ha sido eliminada.")
        self.object.delete()
        return redirect(success_url)
    


from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from .models import DiagramaProyecto
# Importar WeasyPrint si no está ya
try:
    from weasyprint import HTML
except ImportError:
    HTML = None


# ✅ NUEVA VISTA: Para renderizar la página del editor de diagramas
class DiagramaEditorView(LoginRequiredMixin, TemplateView):
    template_name = 'ventas/diagram_editor.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'pk' in kwargs:
            # Estamos editando un diagrama existente
            diagrama = get_object_or_404(DiagramaProyecto, pk=kwargs['pk'])
            context['diagrama'] = diagrama
            context['proyecto'] = diagrama.proyecto
        elif 'proyecto_pk' in kwargs:
            # Estamos creando un nuevo diagrama para un proyecto
            proyecto = get_object_or_404(Proyecto, pk=kwargs['proyecto_pk'])
            context['proyecto'] = proyecto
        return context

# ✅ NUEVA VISTA API: Para devolver los datos JSON de un diagrama
@login_required
def get_diagrama_api(request, diagrama_pk):
    diagrama = get_object_or_404(DiagramaProyecto, pk=diagrama_pk)
    return JsonResponse({
        'id': diagrama.id,
        'titulo': diagrama.titulo,
        # Aquí se decodifica el string de la BD a un objeto JSON para el cliente
        'codigo': json.loads(diagrama.codigo) 
    })

# ✏️ VISTA MODIFICADA: Para guardar el JSON y el SVG del diagrama
@login_required
def guardar_diagrama_api(request, proyecto_pk):
    """
    API para crear o actualizar un diagrama desde el editor JointJS.
    Esta es la versión corregida y definitiva.
    """
    proyecto = get_object_or_404(Proyecto, pk=proyecto_pk)
    
    if request.method == 'POST':
        try:
            # 1. Parseamos el cuerpo del request, que es un JSON principal
            data = json.loads(request.body)
            
            diagrama_id = data.get('id')
            titulo = data.get('titulo', 'Diagrama sin título')
            
            # 2. Extraemos el 'codigo'. Su valor es un STRING que contiene el JSON del diagrama.
            #    Esto es correcto porque el frontend ya hizo JSON.stringify() sobre el objeto del grafo.
            codigo_json_string = data.get('codigo', '{}')
            
            # 3. Extraemos la representación SVG, que también es un string.
            svg_code = data.get('svg', '')

            # 4. Usamos update_or_create para manejar creación y actualización.
            #    Guardamos 'codigo_json_string' directamente en el TextField del modelo.
            diagrama, created = DiagramaProyecto.objects.update_or_create(
                id=diagrama_id,
                defaults={
                    'proyecto': proyecto, 
                    'titulo': titulo, 
                    'codigo': codigo_json_string,
                    'svg_representation': svg_code
                }
            )
            
            # 5. Devolvemos una respuesta exitosa con el ID del diagrama.
            return JsonResponse({'status': 'success', 'diagrama_id': diagrama.id})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'JSON inválido en el request.'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

@login_required
def descargar_diagrama_pdf(request, diagrama_pk):
    """Genera un PDF a partir del SVG guardado de un diagrama."""
    if HTML is None:
        return HttpResponse("WeasyPrint no está instalado.", status=501)
        
    diagrama = get_object_or_404(DiagramaProyecto, pk=diagrama_pk)
    
    # Renderizamos una plantilla que simplemente incrusta el SVG
    html_string = render_to_string('ventas/pdf/diagrama_pdf_template.html', {
        'diagrama': diagrama
    })
    
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="diagrama_{diagrama.titulo}.pdf"'
    
    return response

from django.db import transaction

def reordenar_columnas_api(request, proyecto_pk):
    if request.method == 'POST':
        try:
            # Obtenemos la lista de IDs de las columnas en el nuevo orden
            ordered_ids = json.loads(request.body).get('orden_columnas', [])
            
            with transaction.atomic():
                for index, columna_id in enumerate(ordered_ids):
                    KanbanColumna.objects.filter(id=columna_id, proyecto_id=proyecto_pk).update(orden=index)
            
            return JsonResponse({'status': 'success', 'message': 'Orden de columnas actualizado.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
            
    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)